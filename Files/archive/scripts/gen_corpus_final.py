#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PATH = "/Users/harishsharma/Claude/Pharma BMT/AI_Stress_Test_Corpus.xlsx"
wb = openpyxl.load_workbook(PATH)
ws = wb["Stress Test Messages"]
R = [
 ("Lead Creation","exhibition","trade fair pe mila tha dr kapoor, kapoor clinic rohtak 9898334455, lead daal do","Hinglish","create_lead","name+firm+phone+city","Easy","create_lead full extraction"),
 ("Lead Creation","cold on create","naya lead sharma medical shahbad 9898445566, abhi interest kam hai cold rakhna","Hinglish","create_lead","temp=Cold","Medium","create_lead with temp=Cold"),
 ("Lead Search","lost month","pichle mahine lost hue leads dikhao","Hinglish","search_leads","stage=Lost","Medium","stage=Lost (month range optional)"),
 ("Followups","tom evening","verma ji ka followup kal shaam 6 baje laga do","Hinglish","set_followup","date=tomorrow","Easy","tomorrow resolved"),
 ("Call Logging","ringing","call kiya om medical ko, ringing ho rahi thi kisi ne nahi uthaya","Hinglish","log_call","fu_status=No Answer","Easy","No Answer"),
 ("Stage/Temp Updates","cold3","sanjay medicos ko cold mark karo 6 mahine se kuch nahi hua","Hinglish","update_temp","temp=Cold","Easy","Cold"),
 ("Day Planning","day after plan","parso ka plan batao","Hinglish","get_today_plan","date=+2","Medium","day-after-tomorrow resolved"),
 ("Stats","lost month2","is mahine kitne leads gaye haath se","Hinglish","get_stats","metric=lost, this_month","Medium","lost/this_month"),
 ("Party Management","details4","om pharmacy ka drug license number batao","Hinglish","get_party_details","party=om pharmacy","Medium","get_party_details"),
 ("Party Notes","note5","note likho city drug house pe: new godown khula hai ring road pe","Hinglish","add_party_note","note","Easy","add_party_note"),
 ("Products","details5","zerodol-p ka pack size kya hai","Hinglish","get_product_details","product=zerodol-p","Easy","get_product_details"),
 ("Orders","status3","sharma wala order deliver hua ya nahi","Hinglish","get_order_status","order_query=sharma","Easy","get_order_status by party"),
 ("Dues & Payments","pay5","cheque no 112233 se 22000 aaya krishna drugs se","Hinglish","log_payment","amount=22000, mode=cheque, ref=112233","Easy","log_payment with ref"),
 ("Stock","stock4","sinarest ka stock kitna bacha hai","Hinglish","get_stock_on_hand","product=sinarest","Easy","get_stock_on_hand"),
 ("Navigation","nav4","transporters page pe jao","Hinglish","navigate_to","page=transporters","Easy","navigate_to transporters"),
 ("Transporters","info9","shree shyam transport ka gstin batao","Hinglish","get_transporter_info","transporter=shree shyam transport","Easy","get_transporter_info"),
 ("App Help","help5","lead delete karne ka option kahan hai","Hinglish","app_help","topic=lead delete","Medium","app_help (asking where feature lives)"),
 ("Smalltalk","bye2","ok bye good night","English","smalltalk","none","Easy","smalltalk"),
 ("Unsupported/Refusals","photo edit","meri photo edit kar do profile wali","Hinglish","unsupported","off-task","Easy","unsupported"),
 ("Adversarial & Edge","xss2","<img src=x onerror=alert(1)> sharma ka due","English","get_party_dues","party=sharma","Evil","HTML ignored; get_party_dues(sharma) or unsupported"),
]
start = ws.max_row
fill_a = PatternFill("solid",fgColor="FFFFFF"); fill_b = PatternFill("solid",fgColor="EDEDED")
thin = Side(style="thin",color="C0C0C0"); border = Border(left=thin,right=thin,top=thin,bottom=thin)
diff_color = {"Easy":"1A7A1A","Medium":"B07A00","Hard":"B03A00","Evil":"A00000"}
for i,row in enumerate(R,1):
    idx = start - 1 + i
    vals = [f"ST-{idx:04d}"]+list(row)
    r = start + i
    for c_i,v in enumerate(vals,1):
        cell = ws.cell(row=r,column=c_i,value=v)
        cell.font = Font(name="Calibri",size=10)
        cell.alignment = Alignment(vertical="top",wrap_text=(c_i in (4,7,9)))
        cell.border = border
        cell.fill = fill_a if idx%2 else fill_b
        if c_i==8:
            cell.font = Font(name="Calibri",size=10,bold=True,color=diff_color.get(v,"000000"))
            cell.alignment = Alignment(horizontal="center",vertical="top")
ws.auto_filter.ref = f"A1:I{ws.max_row}"
wb.save(PATH)
wb2 = openpyxl.load_workbook(PATH); w = wb2["Stress Test Messages"]
n = w.max_row-1
empty = sum(1 for r in range(2,w.max_row+1) if not w.cell(row=r,column=6).value)
msgs = set(w.cell(row=r,column=4).value for r in range(2,w.max_row+1))
ids = set(w.cell(row=r,column=1).value for r in range(2,w.max_row+1))
print("rows:",n,"empty intents:",empty,"distinct msgs:",len(msgs),"distinct ids:",len(ids),"sheets:",wb2.sheetnames)
