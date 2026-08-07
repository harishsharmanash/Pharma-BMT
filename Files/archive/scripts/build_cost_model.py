from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()

BLUE = Font(color="0066CC")
BLACK = Font(color="000000")
GREEN = Font(color="1E7D32")
RED = Font(color="CC0000")
HDR_FILL = PatternFill("solid", start_color="122B49")
HDR_FONT = Font(bold=True, color="FFFFFF")
SEC_FONT = Font(bold=True, size=12, color="122B49")
TITLE_FONT = Font(bold=True, size=18, color="122B49")
GRAY = Font(color="777777", size=10)
INR = '"₹"#,##0.00'
INR0 = '"₹"#,##0'
PCT = '0.0%'

def sheet_setup(ws, widths):
    ws.sheet_view.showGridLines = False
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 30

def hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = HDR_FILL; c.font = HDR_FONT
    return c

# ============ 1. COVER ============
ws = wb.active; ws.title = "Cover"
sheet_setup(ws, {"A": 3, "B": 40, "C": 22, "D": 22, "E": 22, "F": 14})
ws.merge_cells("B2:E2")
ws["B2"] = "Acrowell AI Assistant — Cost & Pricing Model"; ws["B2"].font = TITLE_FONT
ws["B2"].alignment = Alignment(horizontal="center")
ws["B4"] = "Editable cost model: Gemini-only vs Hybrid (Qwen3-30B text + Gemma-4-26B vision). Change inputs on the Inputs sheet — everything recalculates."
ws["B4"].font = Font(size=11, color="555555")

ws["B6"] = "Key results (at default inputs)"; ws["B6"].font = SEC_FONT
for i, (label, formula) in enumerate([
    ("All-Gemini cost / user / month", "='Cost Model'!C14"),
    ("Hybrid cost / user / month", "='Cost Model'!D14"),
    ("50-rep company / month (Hybrid)", "='Cost Model'!D16"),
    ("Monthly savings vs Gemini (50 reps)", "='Cost Model'!E18"),
    ("Margin at list price (Hybrid)", "='Cost Model'!D21"),
]):
    r = 7 + i
    ws.cell(row=r, column=2, value=label)
    c = ws.cell(row=r, column=3, value=formula); c.font = GREEN
    c.number_format = PCT if "Margin" in label else INR0
    ws.cell(row=r, column=3).font = GREEN

ws["B14"] = "Sheet index"; ws["B14"].font = SEC_FONT
for i, (name, desc) in enumerate([
    ("Inputs", "All editable assumptions (blue cells) — usage, tokens, model rates, pricing"),
    ("Cost Model", "Per-message → per-user → per-company costs, margins, free-tier offset"),
    ("Scale", "Cost / revenue / profit from 10 to 500 reps + 12-month projection"),
    ("Sources", "Pricing data sources (July 2026)"),
]):
    r = 15 + i
    ws.cell(row=r, column=2, value=name).font = Font(bold=True)
    ws.cell(row=r, column=3, value=desc)
ws["B21"] = "Blue = input you can edit · Black = formula · Green = cross-sheet reference · Red = external data"
ws["B21"].font = GRAY

# ============ 2. INPUTS ============
ws = wb.create_sheet("Inputs")
sheet_setup(ws, {"A": 3, "B": 44, "C": 16, "D": 30})
ws.merge_cells("B2:D2"); ws["B2"] = "Inputs & Assumptions (edit blue cells)"; ws["B2"].font = TITLE_FONT

def put(row, label, value, note="", fmt=None):
    ws.cell(row=row, column=2, value=label)
    c = ws.cell(row=row, column=3, value=value); c.font = BLUE
    if fmt: c.number_format = fmt
    if note: ws.cell(row=row, column=4, value=note).font = GRAY

ws["B4"] = "Usage"; ws["B4"].font = SEC_FONT
put(5, "Messages per user per day", 400)
put(6, "Share of messages that are bill PDFs", 0.5, "order extraction via photo/PDF", PCT)
put(7, "Reps (users)", 50)
put(8, "Billing days per month", 30)

ws["B10"] = "Token profile per message"; ws["B10"].font = SEC_FONT
put(11, "Cached prefix tokens (system prompt + 31 tools)", 2360)
put(12, "Fresh input tokens — text message", 150, "user msg + history + vocab")
put(13, "Output tokens — text intent JSON", 80)
put(14, "Image tokens per bill PDF page", 1000, "after 1024px downscale")
put(15, "Output tokens — PDF order JSON", 400)

ws["B17"] = "Model rates ($ per 1M tokens)"; ws["B17"].font = SEC_FONT
for i, h in enumerate(["Model", "Input $/M", "Cached in $/M", "Output $/M"]):
    ws.cell(row=18, column=2 + i, value=h).font = Font(bold=True)
put(19, "Gemini 3.1 Flash-Lite", 0.25); ws.cell(row=19, column=4, value=0.025).font = BLUE
ws.cell(row=19, column=5, value=1.50).font = BLUE
put(20, "Qwen3-30B-A3B (Workers AI)", 0.051); ws.cell(row=20, column=4, value="—")
ws.cell(row=20, column=5, value=0.335).font = BLUE
put(21, "Gemma-4-26B (Workers AI, vision)", 0.10); ws.cell(row=21, column=4, value="—")
ws.cell(row=21, column=5, value=0.30).font = BLUE

ws["B23"] = "Cloudflare free tier"; ws["B23"].font = SEC_FONT
put(24, "Free neurons per day", 14000)
put(25, "Qwen neurons per text msg", 14, "4625/M in + 30475/M out")
put(26, "Gemma neurons per PDF msg", 43, "9091/M in + 27273/M out")

ws["B28"] = "Pricing & FX"; ws["B28"].font = SEC_FONT
put(29, "List price per user per month (₹)", 999, fmt=INR0)
put(30, "INR per USD", 90)

# ============ 3. COST MODEL ============
ws = wb.create_sheet("Cost Model")
sheet_setup(ws, {"A": 3, "B": 46, "C": 18, "D": 18, "E": 18})
ws.merge_cells("B2:E2"); ws["B2"] = "Cost Model — Gemini-only vs Hybrid"; ws["B2"].font = TITLE_FONT
for i, h in enumerate(["Metric", "All-Gemini", "Hybrid (Qwen+Gemma)", "Notes"]):
    hdr(ws, 4, 2 + i, h)

I = "Inputs!"
# per-message costs in ₹
ws["B5"] = "Cost per text message (₹)"
ws["C5"] = f"=({I}C11*{I}C19*0.1+{I}C12*{I}C19+{I}C13*{I}E19)/1000000*{I}C30"
# note: cached rate is input*0.1 → use D19 instead
ws["C5"] = f"=({I}C11*{I}D19+{I}C12*{I}C19+{I}C13*{I}E19)/1000000*{I}C30"
ws["D5"] = f"=(({I}C11+{I}C12)*{I}C20+{I}C13*{I}E20)/1000000*{I}C30"
ws["E5"] = "cached prefix at 10% rate on Gemini"
ws["B6"] = "Cost per PDF order message (₹)"
ws["C6"] = f"=({I}C11*{I}D19+({I}C12+{I}C14)*{I}C19+{I}C15*{I}E19)/1000000*{I}C30"
ws["D6"] = f"=(({I}C11+{I}C12+{I}C14)*{I}C21+{I}C15*{I}E21)/1000000*{I}C30"
ws["E6"] = "image tokens never cached"

ws["B8"] = "Per user per day (₹)"; ws["B8"].font = SEC_FONT
ws["B9"] = "Text messages (₹)"
ws["C9"] = f"=C5*{I}C5*(1-{I}C6)"; ws["D9"] = f"=D5*{I}C5*(1-{I}C6)"
ws["B10"] = "PDF messages (₹)"
ws["C10"] = f"=C6*{I}C5*{I}C6"; ws["D10"] = f"=D6*{I}C5*{I}C6"
ws["B11"] = "Free-tier offset (₹)"; ws["C11"] = 0
ws["D11"] = f"=-MIN(1,{I}C24/({I}C7*({I}C5*(1-{I}C6)*{I}C25+{I}C5*{I}C6*{I}C26)))*(D9+D10)"
ws["E11"] = "free neurons cover this share of daily use"
ws["B12"] = "Total per user per day (₹)"; ws["B12"].font = Font(bold=True)
ws["C12"] = "=C9+C10+C11"; ws["D12"] = "=D9+D10+D11"
ws["B14"] = "Per user per month (₹)"; ws["B14"].font = Font(bold=True)
ws["C14"] = f"=C12*{I}C8"; ws["D14"] = f"=D12*{I}C8"

ws["B16"] = "Company per month (₹)"; ws["B16"].font = Font(bold=True)
ws["C16"] = f"=C14*{I}C7"; ws["D16"] = f"=D14*{I}C7"
ws["B18"] = "Savings vs Gemini / month (₹)"; ws["E18"] = "=C16-D16"; ws["B18"].font = Font(bold=True)
ws["B19"] = "Savings %"; ws["E19"] = "=E18/C16"; ws["E19"].number_format = PCT

ws["B21"] = "Margin at list price (Hybrid)"; ws["D21"] = f"=({I}C29-D14)/{I}C29"; ws["D21"].number_format = PCT
ws["B22"] = "Profit per company / month (₹)"
ws["C22"] = f"=({I}C29-C14)*{I}C7"; ws["D22"] = f"=({I}C29-D14)*{I}C7"

for r in [5,6,9,10,11,12]:
    for col in "CD": ws[f"{col}{r}"].number_format = INR
for r in [14,16,18,22]:
    for col in "CDE":
        if ws[f"{col}{r}"].value is not None: ws[f"{col}{r}"].number_format = INR0
for r in range(5, 23):
    for col in "CDE":
        c = ws[f"{col}{r}"]
        if isinstance(c.value, str) and c.value.startswith("=") and c.font.color is None:
            c.font = BLACK

wb.save("/Users/harishsharma/Claude/Pharma BMT/AI_Cost_Model.xlsx")
print("sheets 1-3 done")

# ============ 4. SCALE ============
ws = wb.create_sheet("Scale")
sheet_setup(ws, {"A": 3, "B": 14, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})
ws.merge_cells("B2:G2"); ws["B2"] = "Scale — monthly cost, revenue, profit by company size"; ws["B2"].font = TITLE_FONT
for i, h in enumerate(["Reps", "Gemini cost ₹/mo", "Hybrid cost ₹/mo", "Revenue ₹/mo", "Profit (Hybrid) ₹/mo", "Margin %"]):
    hdr(ws, 4, 2 + i, h)
for j, reps in enumerate([10, 25, 50, 100, 250, 500]):
    r = 5 + j
    ws.cell(row=r, column=2, value=reps).font = BLUE
    ws.cell(row=r, column=3, value=f"='Cost Model'!C14*B{r}").number_format = INR0
    ws.cell(row=r, column=4, value=f"='Cost Model'!D14*B{r}").number_format = INR0
    ws.cell(row=r, column=5, value=f"=Inputs!C29*B{r}").number_format = INR0
    ws.cell(row=r, column=6, value=f"=E{r}-D{r}").number_format = INR0
    ws.cell(row=r, column=7, value=f"=F{r}/E{r}").number_format = PCT
    ws.cell(row=r, column=3).font = GREEN; ws.cell(row=r, column=4).font = GREEN

chart = BarChart(); chart.type = "col"; chart.title = "Monthly AI cost: Gemini vs Hybrid"
data = Reference(ws, min_col=3, max_col=4, min_row=4, max_row=10)
cats = Reference(ws, min_col=2, min_row=5, max_row=10)
chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
chart.height = 8; chart.width = 16
ws.add_chart(chart, "B13")

ws["B32"] = "12-month projection (Hybrid, starting reps, +15% growth/mo)"; ws["B32"].font = SEC_FONT
for i, h in enumerate(["Month", "Reps", "Cost ₹", "Revenue ₹", "Profit ₹"]):
    hdr(ws, 33, 2 + i, h)
for m in range(1, 13):
    r = 33 + m
    ws.cell(row=r, column=2, value=m)
    if m == 1:
        ws.cell(row=r, column=3, value="=Inputs!C7").font = GREEN
    else:
        ws.cell(row=r, column=3, value=f"=ROUND(C{r-1}*1.15,0)")
    ws.cell(row=r, column=4, value=f"='Cost Model'!D14*C{r}").number_format = INR0
    ws.cell(row=r, column=5, value=f"=Inputs!C29*C{r}").number_format = INR0
    ws.cell(row=r, column=6, value=f"=E{r}-D{r}").number_format = INR0
    ws.cell(row=r, column=4).font = GREEN
lchart = BarChart(); lchart.type = "col"; lchart.title = "Profit per month (Hybrid, 15% growth)"
ldata = Reference(ws, min_col=6, min_row=33, max_row=45)
lcats = Reference(ws, min_col=2, min_row=34, max_row=45)
lchart.add_data(ldata, titles_from_data=True); lchart.set_categories(lcats)
lchart.height = 8; lchart.width = 16; lchart.legend = None
ws.add_chart(lchart, "B48")

wb.save("/Users/harishsharma/Claude/Pharma BMT/AI_Cost_Model.xlsx")
print("scale done")

# ============ 5. SOURCES ============
ws = wb.create_sheet("Sources")
sheet_setup(ws, {"A": 3, "B": 40, "C": 30, "D": 60})
ws.merge_cells("B2:D2"); ws["B2"] = "Sources (pricing verified July 2026)"; ws["B2"].font = TITLE_FONT
for i, h in enumerate(["Data", "Source Name", "Source URL"]):
    hdr(ws, 4, 2 + i, h)
rows = [
    ("Gemini 3.1 Flash-Lite $0.25/$1.50 per M, cache ~10%", "BenchLM Gemini pricing (synced from Google)", "https://benchlm.ai/google/api-pricing"),
    ("Gemini caching 90% savings + Batch 50% off", "CostGoat Gemini guide", "https://costgoat.com/pricing/gemini-api"),
    ("Qwen3-30B / Gemma-4-26B / neuron rates", "Cloudflare Workers AI pricing docs", "https://developers.cloudflare.com/workers-ai/platform/pricing/"),
    ("Free neuron allocation (~10K/day)", "Cloudflare Workers AI docs", "https://developers.cloudflare.com/workers-ai/platform/pricing/"),
    ("LLM market pricing context", "Opslyft LLM pricing 2026", "https://www.opslyft.com/blog/llm-pricing-comparison-cost-per-token"),
]
for j, row in enumerate(rows):
    r = 5 + j
    for k, v in enumerate(row):
        c = ws.cell(row=r, column=2 + k, value=v); c.font = RED
ws["B11"] = "Note: re-verify rates on official pages before contract decisions; third-party trackers were used."
ws["B11"].font = GRAY

wb.save("/Users/harishsharma/Claude/Pharma BMT/AI_Cost_Model.xlsx")
print("all sheets saved")
